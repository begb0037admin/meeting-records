from __future__ import annotations

import base64
import io
import json
import re
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import PurePosixPath
from typing import Any
from urllib.parse import quote


OWNER = "begb0037admin"
SOURCES = {
    "inbox": ("work-inbox", "data/briefing.json"),
    "tasks": ("command-centre", "data/tasks.json"),
    "roadmap": ("hr-projects", "HR Systems Roadmap/HR Systems Roadmap MASTER.xlsm"),
    "meeting": ("meeting-records", "Meeting Reviews/HR Systems Managers Meeting — 24-06.md"),
}
SAFE_REF = re.compile(r"^[A-Za-z0-9._/-]{1,200}$")
REFRESH_CONFIRMATION = "REFRESH OUTLOOK AND PUBLISH WORK INBOX"


class SourceError(RuntimeError):
    pass


def require_refresh_confirmation(confirmation: str, *, enabled: bool = False) -> None:
    if not enabled:
        raise SourceError(
            "Outlook refresh is disabled for this MCP session. Start a separately "
            "authorised refresh invocation; normal meeting preparation cannot enable it."
        )
    if confirmation != REFRESH_CONFIRMATION:
        raise SourceError(
            "Outlook refresh was not run. To perform the explicit refresh, pass the exact "
            f"confirmation: {REFRESH_CONFIRMATION}"
        )


@dataclass(frozen=True)
class GitHubBlob:
    repo: str
    path: str
    ref: str
    sha: str
    content: bytes


class GitHubReader:
    """GitHub-only reader. Every call is an HTTP GET through authenticated gh."""

    def _api(self, route: str) -> Any:
        completed = subprocess.run(
            ["gh", "api", "--method", "GET", route],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        if completed.returncode:
            raise SourceError(completed.stderr.strip() or f"GitHub read failed: {route}")
        try:
            return json.loads(completed.stdout)
        except json.JSONDecodeError as exc:
            raise SourceError(f"GitHub returned invalid JSON for {route}") from exc

    def blob(self, repo: str, path: str, ref: str = "main") -> GitHubBlob:
        if not SAFE_REF.fullmatch(ref):
            raise SourceError("Invalid Git reference")
        encoded_path = quote(str(PurePosixPath(path)), safe="/")
        payload = self._api(f"repos/{OWNER}/{repo}/contents/{encoded_path}?ref={quote(ref, safe='')}")
        if payload.get("type") != "file":
            raise SourceError(f"{repo}/{path} is not a file")
        return GitHubBlob(
            repo=repo,
            path=path,
            ref=ref,
            sha=payload["sha"],
            content=base64.b64decode(payload["content"]),
        )

    def json_file(self, repo: str, path: str, ref: str = "main") -> tuple[Any, GitHubBlob]:
        blob = self.blob(repo, path, ref)
        return json.loads(blob.content.decode("utf-8-sig")), blob


def compact_briefing(payload: dict[str, Any], blob: GitHubBlob) -> dict[str, Any]:
    sections: dict[str, list[dict[str, Any]]] = {}
    for key in ("urgent", "needs", "waiting", "fyi"):
        values = payload.get(key, [])
        if isinstance(values, list):
            sections[key] = [
                {
                    "title": item.get("title") or item.get("subject"),
                    "from": item.get("from"),
                    "received": item.get("received"),
                    "badge": item.get("badge"),
                    "summary": item.get("sub"),
                }
                for item in values
            ]
    return {
        "source": f"github:{OWNER}/{blob.repo}/{blob.path}@{blob.ref}",
        "blobSha": blob.sha,
        "snapshotDate": payload.get("date"),
        "subtitle": payload.get("subtitle"),
        "context": payload.get("context"),
        "sections": sections,
        "readOnly": True,
        "outlookRefreshed": False,
    }


def filter_tasks(
    tasks: list[dict[str, Any]],
    *,
    query: str = "",
    tiers: list[str] | None = None,
    include_done: bool = False,
    limit: int = 50,
) -> list[dict[str, Any]]:
    needle = query.casefold().strip()
    tier_set = {item.casefold() for item in tiers or []}
    result = []
    for task in tasks:
        if not include_done and (task.get("done") is True or str(task.get("status", "")).casefold() == "done"):
            continue
        if tier_set and str(task.get("tier", "")).casefold() not in tier_set:
            continue
        searchable = " ".join(
            str(task.get(key, "")) for key in ("id", "title", "summary", "description", "source", "emailRef")
        ).casefold()
        if needle and needle not in searchable:
            continue
        result.append({key: task.get(key) for key in (
            "id", "title", "tier", "source", "emailRef", "summary", "description",
            "actions", "dateAdded", "status", "notes",
        )})
        if len(result) >= max(1, min(limit, 200)):
            break
    return result


def _normalise_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def read_roadmap(
    content: bytes,
    *,
    query: str = "",
    statuses: list[str] | None = None,
    lead: str = "",
    limit: int = 50,
) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    stream = io.BytesIO(content)
    workbook = load_workbook(stream, read_only=True, data_only=True, keep_vba=False)
    try:
        sheet = workbook["Work Tracker"]
        fields = {
            "id": 0, "activity": 1, "lead": 6, "team": 7, "deadline": 19,
            "progress": 22, "nextSteps": 23, "lastReviewed": 24,
            "status": 26, "statusDetails": 27,
        }
        needle = query.casefold().strip()
        status_set = {item.casefold() for item in statuses or []}
        lead_needle = lead.casefold().strip()
        results: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not row[fields["activity"]]:
                continue
            item = {name: _normalise_cell(row[index] if index < len(row) else None) for name, index in fields.items()}
            if status_set and str(item["status"] or "").casefold() not in status_set:
                continue
            if lead_needle and lead_needle not in str(item["lead"] or "").casefold():
                continue
            if needle and needle not in " ".join(str(value or "") for value in item.values()).casefold():
                continue
            results.append(item)
            if len(results) >= max(1, min(limit, 200)):
                break
        return results
    finally:
        workbook.close()
        stream.close()



def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()
